from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import FileResponse
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import os

app = FastAPI(
    title="地址資料整理服務",
    description="上傳 CSV 檔案，進行地址篩選與標黃處理。"
)

# 你的資料處理核心邏輯，包裝成一個函數
def process_excel_highlight(df: pd.DataFrame, column_name: str) -> BytesIO:
    """處理 DataFrame 並返回標黃後的 Excel BytesIO 物件"""
    
    # 篩選邏輯
    regex_pattern = r"([a-zA-Z].*[a-zA-Z].*[a-zA-Z])|(\d{4,})"
    mask = df[column_name].astype(str).apply(
        lambda x: bool(re.search(regex_pattern, x, re.IGNORECASE))
    )

    # 1. 寫入 Pandas 到 BytesIO (作為 Excel 基礎)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='標黃結果')
    
    output.seek(0)
    
    # 2. OpenPyXL 上色
    wb = load_workbook(output)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for idx, flag in enumerate(mask, start=2):
        if flag:
            for cell in ws[idx]:
                cell.fill = yellow_fill

    # 3. 儲存結果到新的 BytesIO
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output

# API 接口：處理檔案上傳
@app.post("/process_file/")
async def create_upload_file(
    file: UploadFile = File(...), 
    column_name: str = Form("住所２")
):
    # 讀取上傳的 CSV 內容
    contents = await file.read()
    try:
        df = pd.read_csv(BytesIO(contents))
    except Exception as e:
        return {"message": "檔案讀取失敗，請確認檔案格式是否為有效的 CSV。", "error": str(e)}

    # 檢查欄位
    if column_name not in df.columns:
        return {"message": f"錯誤：找不到指定的欄位『{column_name}』。"}

    # 執行處理
    excel_buffer = process_excel_highlight(df, column_name)
    
    # 儲存到暫存檔案，以便 FileResponse 返回
    temp_filename = f"processed_{file.filename.split('.')[0]}.xlsx"
    with open(temp_filename, "wb") as f:
        f.write(excel_buffer.getbuffer())

    # 返回 Excel 檔案供下載
    return FileResponse(
        temp_filename, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=f"整理結果_{file.filename.split('.')[0]}.xlsx"
    )

# 運行說明：
# 需要安裝 uvicorn: pip install uvicorn
# 運行指令: uvicorn main:app --reload

