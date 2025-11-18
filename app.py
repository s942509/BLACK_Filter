import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

# 設定頁面標題
st.set_page_config(page_title="CSV資料篩選與標黃工具", layout="wide")
st.title("📝 數據清洗與地址標記工具")

# --- 資料處理主函數 ---
@st.cache_data
def process_data(uploaded_file, column_name="住所２"):
    """
    接收上傳的 CSV 檔案，進行篩選與標黃處理。
    返回一個 BytesIO 物件，其中包含標黃後的 Excel 檔案內容。
    """
    try:
        # 1. 讀取 CSV
        df = pd.read_csv(uploaded_file)
        
        # 檢查欄位是否存在
        if column_name not in df.columns:
            st.error(f"⚠️ 錯誤：上傳的 CSV 檔案中找不到指定的欄位『{column_name}』。請確認檔案內容或更改欄位名稱。")
            return None, None

        # 2. 找出符合條件的列 (篩選邏輯)
        # 條件：至少包含三個英文字母 OR 包含四個或更多連續數字
        regex_pattern = r"([a-zA-Z].*[a-zA-Z].*[a-zA-Z])|(\d{4,})"
        mask = df[column_name].astype(str).apply(
            lambda x: bool(re.search(regex_pattern, x, re.IGNORECASE))
        )

        # 3. 準備 Excel 寫入的緩衝區
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='標黃結果')
        writer.close()
        
        # 4. 打開 Excel 並上色
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 從第二列開始遍歷 (第一列是表頭)
        for idx, flag in enumerate(mask, start=2):
            if flag:
                # 遍歷該列的所有儲存格並上色
                for cell in ws[idx]:
                    cell.fill = yellow_fill

        # 5. 將結果存入 BytesIO 物件
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        # 顯示標記了多少筆資料
        marked_count = mask.sum()
        
        return final_output, marked_count

    except Exception as e:
        st.error(f"處理檔案時發生錯誤：{e}")
        return None, None

# --- Streamlit UI 介面 ---

# 讓使用者輸入要檢查的欄位名稱 (預設為 '住所２')
column_to_check = st.sidebar.text_input(
    "👉 請輸入要檢查的欄位名稱：", 
    value="住所２",
    help="這是您 CSV 檔案中包含地址或需要檢查文字/數字的欄位。"
)

# 檔案上傳元件
uploaded_file = st.file_uploader(
    "上傳您的 CSV 檔案 (.csv)", 
    type=["csv"],
    help="上傳後，程式將自動執行篩選並生成帶有標黃結果的 Excel 檔案。"
)

st.markdown("""
### 篩選條件說明：
程式將標記符合以下任一條件的列（整列標黃）：
1.  **至少包含三個英文字母** (例如：ABC, A.B.C, TPE)
2.  **包含四個或更多連續數字** (例如：1234, 56789)
""")

if uploaded_file is not None:
    st.info(f"檔案已上傳。正在對 **『{column_to_check}』** 欄位執行處理...")
    
    # 呼叫處理函數
    excel_buffer, count = process_data(uploaded_file, column_to_check)

    if excel_buffer:
        st.success(f"🎉 處理完成！共標記了 **{count}** 筆符合條件的資料。")
        
        # 下載按鈕
        st.download_button(
            label="⬇️ 點此下載結果 (Excel)",
            data=excel_buffer,
            file_name="整理結果_標黃.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="下載包含標黃結果的 Excel 檔案。"
        )
        
        # 顯示處理後 DataFrame 的前幾行預覽
        st.subheader("📁 資料預覽 (前 5 行)")
        uploaded_file.seek(0) # 重設檔案指標以供再次讀取
        preview_df = pd.read_csv(uploaded_file)
        st.dataframe(preview_df.head())
